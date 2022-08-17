# Creator: Reenphy George
# Date : 26th December
# Title: Sheets to VCF with batch rename

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
a=(input("Enter the name of file : "))
wb = load_workbook(a)
ws = wb.active

# function to group rename
def rename():
    str_append = input("Enter the string to be appended: ")

    name = 1
    while (ws[get_column_letter(name) + '1'].value).lower() != 'name':
        name += 1

    row = ws.max_row
    for i in range(2,row+1):
        exl = ws[get_column_letter(name) + str(i)].value
        exl = exl + " " + str_append
        ws[get_column_letter(name) + str(i)].value = exl

    wb.save('sheet.xlsx')
    print("Successfully Renamed ")
    return name

# function to create vcf
def vcf(name):
    contact = 1
    while (ws[get_column_letter(contact) + '1'].value).lower() != 'phone':
        contact += 1
    
    row = ws.max_row
    f = open('temp.txt' , 'a')
    for i in range(2,row+1):
        fn = str(ws[get_column_letter(name) + str(i)].value)
        tel = str(ws[get_column_letter(contact) + str(i)].value)
        f.write("BEGIN:VCARD\nFN:"+fn+"\nTEL:"+tel+"\nEND:VCARD\n\n")
    
    os.renames('temp.txt', 'contact.vcf')
    print("VCF Exported Succesfully")
        
name = rename()
vcf(name)
